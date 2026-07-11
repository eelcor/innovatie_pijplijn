"""Excel export routes — exporteer alle data naar een opgemaakt Excel-bestand."""

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import (
    CentralQuestion,
    CurationItem,
    Hypothesis,
    Initiative,
    InitiativeQuestion,
    InitiativeTag,
    Tag,
)

router = APIRouter()

# --- Stijlconstanten ---
_ACCENT_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14, color="0F172A")
_SUBTITLE_FONT = Font(italic=True, size=10, color="64748B")
_BODY_FONT = Font(size=11)
_BORDER_THIN = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
_ALT_ROW_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

# Fasekleuren voor conditional styling
_PHASE_COLORS = {
    "verkenning": ("4F46E5", "EEF2FF"),
    "experiment": ("EA580C", "#FFF7ED"),
    "pilot": ("059669", "ECFDF5"),
    "opschaling": ("2563EB", "EFF6FF"),
}


def _style_header(ws, row: int, col_start: int = 1):
    """Style header row with accent fill and white font."""
    for cell in ws[row]:
        if cell.column >= col_start:
            cell.font = _HEADER_FONT
            cell.fill = _ACCENT_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = _BORDER_THIN


def _style_data_rows(ws, row_start: int, row_end: int):
    """Style data rows with alternating fills and borders."""
    for r in range(row_start, row_end + 1):
        fill = None if (r - row_start) % 2 == 0 else _ALT_ROW_FILL
        for cell in ws[r]:
            cell.font = _BODY_FONT
            cell.border = _BORDER_THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill


def _auto_width(ws, min_w=12, max_w=50):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # Account for multi-line text
                lines = str(cell.value).split('\n')
                longest_line = max(len(l) for l in lines) if lines else 0
                max_length = max(max_length, longest_line + 2)
        adjusted_width = min(max(max_length, min_w), max_w)
        ws.column_dimensions[col_letter].width = adjusted_width


def _add_title(ws, title: str, subtitle: str = ""):
    """Add a styled title row at the top."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = _TITLE_FONT
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ws.max_column)
        cell = ws.cell(row=2, column=1, value=subtitle)
        cell.font = _SUBTITLE_FONT


@router.get("/excel")
async def export_to_excel(request: Request, db: Session = Depends(get_db)):
    """Exporteer alle data naar een opgemaakt Excel-bestand met meerdere tabs."""
    wb = Workbook()

    # ============================================================
    # TAB 1: Initiatieven
    # ============================================================
    ws_init = wb.active
    ws_init.title = "Initiatieven"

    initiatives = db.query(Initiative).order_by(Initiative.title.asc()).all()

    # Haal koppelingen op (vragen en tags per initiatief)
    all_iq = db.query(InitiativeQuestion).all()
    iq_map = {}  # initiative_id -> [question_id]
    for iq in all_iq:
        iq_map.setdefault(iq.initiative_id, []).append(iq.central_question_id)

    all_it = db.query(InitiativeTag).all()
    it_map = {}  # initiative_id -> [tag_id]
    for it in all_it:
        it_map.setdefault(it.initiative_id, []).append(it.tag_id)

    # Haal vragen en tags op voor lookup
    question_map = {q.id: q.question for q in db.query(CentralQuestion).all()}
    tag_map = {t.id: t.name for t in db.query(Tag).filter(Tag.is_active == True).all()}

    # Headers
    headers = [
        "Titel", "Fase", "Status", "Horizon", "MDS", "Eigenaar / team",
        "Trekker", "Centrale vragen", "Tags", "Beschrijving"
    ]
    ws_init.append(headers)

    for i, init in enumerate(initiatives):
        questions = [question_map.get(qid, "") for qid in iq_map.get(init.id, [])]
        tags = [tag_map.get(tid, "") for tid in it_map.get(init.id, [])]

        ws_init.append([
            init.title,
            init.phase,
            init.status,
            init.horizon or "",
            init.mds or "",
            init.owner or "",
            init.trekker or "",
            "; ".join(questions),
            "; ".join(tags),
            init.description or "",
        ])

    _style_header(ws_init, 1)
    _style_data_rows(ws_init, 2, 1 + len(initiatives))
    _auto_width(ws_init)

    # Freeze header row
    ws_init.freeze_panes = "A3"

    # ============================================================
    # TAB 2: Centrale vragen
    # ============================================================
    ws_vr = wb.create_sheet("Centrale vragen")

    questions = db.query(CentralQuestion).filter(
        CentralQuestion.is_active == True
    ).order_by(CentralQuestion.question.asc()).all()

    # Teller initiatieven per vraag
    iq_counts = {}
    for iq in all_iq:
        iq_counts[iq.central_question_id] = iq_counts.get(iq.central_question_id, 0) + 1

    headers_vr = ["Vraag", "Beschrijving", "Aantal initiatieven"]
    ws_vr.append(headers_vr)

    for q in questions:
        ws_vr.append([
            q.question,
            q.description or "",
            iq_counts.get(q.id, 0),
        ])

    _style_header(ws_vr, 1)
    _style_data_rows(ws_vr, 2, 1 + len(questions))
    _auto_width(ws_vr)
    ws_vr.freeze_panes = "A3"

    # ============================================================
    # TAB 3: Hypothesen
    # ============================================================
    ws_hyp = wb.create_sheet("Hypothesen")

    hypotheses = db.query(Hypothesis).order_by(
        Hypothesis.initiative_id.asc(),
        Hypothesis.created_at.asc()
    ).all()

    init_map = {i.id: i.title for i in initiatives}

    headers_hyp = [
        "Initiatief", "Type", "Hypothese", "Leeruitkomst", "Toelichting", "Status", "Datum"
    ]
    ws_hyp.append(headers_hyp)

    for h in hypotheses:
        ws_hyp.append([
            init_map.get(h.initiative_id, ""),
            h.type or "",
            h.description or "",
            h.learning or "",
            h.commentary or "",
            h.status or "",
            h.created_at.strftime("%d-%m-%Y") if h.created_at else "",
        ])

    _style_header(ws_hyp, 1)
    _style_data_rows(ws_hyp, 2, 1 + len(hypotheses))
    _auto_width(ws_hyp)
    ws_hyp.freeze_panes = "A3"

    # ============================================================
    # TAB 4: Curaties
    # ============================================================
    ws_cur = wb.create_sheet("Curaties")

    from app.models import Curation

    curations = db.query(Curation).order_by(Curation.name.asc()).all()

    # Haal items per curatie op
    all_items = db.query(CurationItem).all()
    items_map = {}  # curation_id -> [item]
    for item in all_items:
        items_map.setdefault(item.curation_id, []).append(item)

    headers_cur = ["Naam", "Doel", "Beschrijving", "Aantal initiatieven"]
    ws_cur.append(headers_cur)

    for c in curations:
        item_count = len(items_map.get(c.id, []))
        ws_cur.append([
            c.name,
            c.purpose or "",
            c.description or "",
            item_count,
        ])

    _style_header(ws_cur, 1)
    _style_data_rows(ws_cur, 2, 1 + len(curations))
    _auto_width(ws_cur)
    ws_cur.freeze_panes = "A3"

    # ============================================================
    # TAB 5: Curatie-items (gedetailleerd)
    # ============================================================
    ws_ci = wb.create_sheet("Curatie-items")

    headers_ci = ["Curatie", "Initiatief", "Positie", "Notitie"]
    ws_ci.append(headers_ci)

    curation_map = {c.id: c.name for c in curations}

    for item in all_items:
        ws_ci.append([
            curation_map.get(item.curation_id, ""),
            init_map.get(item.initiative_id, ""),
            item.position or "",
            item.note or "",
        ])

    _style_header(ws_ci, 1)
    _style_data_rows(ws_ci, 2, 1 + len(all_items))
    _auto_width(ws_ci)
    ws_ci.freeze_panes = "A3"

    # ============================================================
    # Stroom de file naar de browser
    # ============================================================
    from io import BytesIO

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="innovatiepijplijn_export.xlsx"',
        },
    )
