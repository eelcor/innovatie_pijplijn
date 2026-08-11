#!/usr/bin/env python3
"""Importeer initiatieven uit Excel v0.2 naar de database.

Leest '20260806 Input/20260806 Inventarisatie_AI_initiatieven_v0.2.xlsx'
en update/aanmaakt initiatieven in de database.

Matching: op ID (bijv. B-01, D-03) of op titel als ID niet matcht.
"""

import sys
import re
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from sqlalchemy.orm import Session
from app.database import SessionLocal, init_db
from app.models import Initiative
from app.logging_config import logger

# --- Mapping tabellen ---

# Excel fase → DB fase (bestaande enum waarden)
FASE_MAP = {
    "idee": "idee",
    "idee (vóór verkenning)": "idee",
    "verkenning": "verkenning",
    "verkenning (passief)": "verkenning",
    "verkenning / experiment": "verkenning",
    "verkenning → start experiment": "verkenning",
    "experiment": "experiment",
    "experiment (gok)": "experiment",
    "experiment (in aanbouw)": "experiment",
    "experiment / opschaling (al in gebruik)": "experiment",
    "pilot": "pilot",
    "opschaling (al in gebruik)": "opschaling",
    "opschaling (alternatieve oplossing in gebruik)": "opschaling",
    "onbekend": "idee",  # fallback
}

# Excel status → DB status
STATUS_MAP = {
    "actief": "actief",
    "gestopt": "gestopt",
    "afgerond": "afgerond",
    "onduidelijk": "actief",  # temporary fallback — status enum moet worden uitgebreid
    "pauze": "actief",        # temporary fallback
    "idee": "actief",         # temporary fallback
    "onbekend": "actief",     # temporary fallback
}

# Excel type AI-gebruik → DB enum
TYPE_AI_MAP = {
    "bouwen met ai": "bouwen_met_ai",
    "ai in bouwsels": "ai_in_bouwsels",
    "ai in bestaande tools": "ai_in_bestaande_tools",
    "persoonlijke productiviteit": "persoonlijke_productiviteit",
    "mix": "mix",
}

# Excel horizon → DB enum
HORIZON_MAP = {
    "h1": "h1",
    "h2": "h2",
    "h3": "h3",
}

# "(niet vermeld)" marker
NIET_VERMELD = "(niet vermeld)"


def normalize(s):
    """Zet '(niet vermeld)' en lege strings naar None."""
    if not s or str(s).strip() == "" or str(s).strip() == NIET_VERMELD:
        return None
    return str(s).strip()


def extract_id(title_or_id):
    """Haal het initiatief ID eruit (bijv. 'B-01' uit '[B-01] Titel')."""
    if not title_or_id:
        return None
    s = str(title_or_id).strip()
    # Match pattern like B-01, D-03, etc.
    m = re.match(r"([A-Z]+-\d+)", s)
    if m:
        return m.group(1)
    return s


def find_existing_initiative(db: Session, excel_id: str, title: str):
    """Zoek bestaand initiatief op ID of titel."""
    if not excel_id:
        return None

    # Probeer eerst op ID in de titel (bestaande records hebben [B-01] prefix)
    init = db.query(Initiative).filter(
        Initiative.title.like(f"[{excel_id}]%")
    ).first()
    if init:
        return init

    # Probeer op exacte ID match in eigen id veld
    init = db.query(Initiative).filter(
        Initiative.id == excel_id
    ).first()
    if init:
        return init

    # Fallback: zoek op titel (zonder ID prefix)
    clean_title = re.sub(r"^\[?[A-Z]+-\d+\]?\s*", "", title).strip()
    init = db.query(Initiative).filter(
        Initiative.title.ilike(f"%{clean_title[:30]}%")
    ).first()
    return init


def import_initiatives(excel_path: str):
    """Importeer alle initiatieven uit het Excel bestand."""
    print(f"Lees Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Initiatieven"]

    # Haal kolom indices op
    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers) if name}

    print(f"Kolommen: {len(headers)} gevonden")
    print(f"Rijen te verwerken: {ws.max_row - 1}")
    print()

    init_db()
    db = SessionLocal()

    stats = {"updated": 0, "created": 0, "errors": 0, "skipped": 0}

    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
            # Haal waarden op met kolom map
            excel_id = normalize(row[col.get("ID", 0)])
            title = normalize(row[col.get("Titel", 4)])
            description = normalize(row[col.get("Omschrijving", 5)])

            if not title:
                stats["skipped"] += 1
                print(f"  Row {row_idx}: geskipped (geen titel)")
                continue

            try:
                # Map waarden
                fase_raw = normalize(row[col.get("Fase", 6)])
                status_raw = normalize(row[col.get("Status", 7)])
                type_ai_raw = normalize(row[col.get("Type AI-gebruik", 8)])
                horizon_raw = normalize(row[col.get("Horizon", 9)])

                phase = FASE_MAP.get(fase_raw, "idee") if fase_raw else None
                status = STATUS_MAP.get(status_raw, "actief") if status_raw else None
                type_ai = TYPE_AI_MAP.get(type_ai_raw.lower(), None) if type_ai_raw else None
                horizon = HORIZON_MAP.get(horizon_raw.lower(), None) if horizon_raw else None

                # Nieuwe velden
                cluster = normalize(row[col.get("Cluster", 1)])
                afdeling = normalize(row[col.get("Afdeling", 2)])
                team = normalize(row[col.get("Team", 3)])
                trekker = normalize(row[col.get("Trekker", 13)])
                potentie = normalize(row[col.get("Potentie", 10)])
                capaciteitsvraag = normalize(row[col.get("Capaciteitsvraag", 11)])
                risico = normalize(row[col.get("Risico", 12)])
                bron_initiatief = normalize(row[col.get("Bron initiatief", 14)])
                externe_partners = normalize(row[col.get("Externe partners", 15)])
                betrokkenheid_iv = normalize(row[col.get("Betrokkenheid IV", 16)])
                centrale_vraag = normalize(row[col.get("Centrale vraag / Aandachtspunten", 17)])
                gerelateerde = normalize(row[col.get("Gerelateerde initiatieven", 18)])
                volgende_stap = normalize(row[col.get("Volgende stap", 19)])
                opmerkingen = normalize(row[col.get("Opmerkingen", 20)])

                # Zoek bestaand initiatief
                existing = find_existing_initiative(db, excel_id, title)

                if existing:
                    # Update bestaand initiatief
                    existing.title = f"[{excel_id}] {title}" if excel_id and not existing.title.startswith(f"[{excel_id}]") else existing.title
                    existing.description = description
                    if phase:
                        existing.phase = phase
                    if status:
                        existing.status = status
                    if type_ai:
                        existing.type_ai_gebruik = type_ai
                    if horizon:
                        existing.horizon = horizon
                    if trekker:
                        existing.trekker = trekker
                    if centrale_vraag:
                        existing.central_question = centrale_vraag

                    # Nieuwe velden (als kolom bestaat)
                    try:
                        if cluster:
                            setattr(existing, "cluster", cluster)
                        if afdeling:
                            setattr(existing, "afdeling", afdeling)
                        if team:
                            setattr(existing, "team", team)
                        if potentie:
                            setattr(existing, "potentie", potentie)
                        if capaciteitsvraag:
                            setattr(existing, "capaciteitsvraag", capaciteitsvraag)
                        if risico:
                            setattr(existing, "risico", risico)
                        if bron_initiatief:
                            setattr(existing, "bron_initiatief", bron_initiatief)
                        if externe_partners:
                            setattr(existing, "externe_partners", externe_partners)
                        if betrokkenheid_iv:
                            setattr(existing, "betrokkenheid_iv", betrokkenheid_iv)
                        if gerelateerde:
                            setattr(existing, "gerelateerde_initiatieven", gerelateerde)
                        if volgende_stap:
                            setattr(existing, "volgende_stap", volgende_stap)
                        if opmerkingen:
                            setattr(existing, "opmerkingen", opmerkingen)
                    except AttributeError:
                        pass  # Kolom bestaat nog niet in model

                    stats["updated"] += 1
                    print(f"  ✓ Updated: [{excel_id}] {title[:50]}")
                else:
                    # Maak nieuw initiatief aan
                    import uuid
                    new_init = Initiative(
                        id=str(uuid.uuid4()),
                        title=f"[{excel_id}] {title}" if excel_id else title,
                        description=description,
                        phase=phase or "idee",
                        status=status or "actief",
                        type_ai_gebruik=type_ai,
                        horizon=horizon,
                        trekker=trekker,
                        central_question=centrale_vraag,
                    )

                    # Nieuwe velden
                    try:
                        if cluster:
                            setattr(new_init, "cluster", cluster)
                        if afdeling:
                            setattr(new_init, "afdeling", afdeling)
                        if team:
                            setattr(new_init, "team", team)
                        if potentie:
                            setattr(new_init, "potentie", potentie)
                        if capaciteitsvraag:
                            setattr(new_init, "capaciteitsvraag", capaciteitsvraag)
                        if risico:
                            setattr(new_init, "risico", risico)
                        if bron_initiatief:
                            setattr(new_init, "bron_initiatief", bron_initiatief)
                        if externe_partners:
                            setattr(new_init, "externe_partners", externe_partners)
                        if betrokkenheid_iv:
                            setattr(new_init, "betrokkenheid_iv", betrokkenheid_iv)
                        if gerelateerde:
                            setattr(new_init, "gerelateerde_initiatieven", gerelateerde)
                        if volgende_stap:
                            setattr(new_init, "volgende_stap", volgende_stap)
                        if opmerkingen:
                            setattr(new_init, "opmerkingen", opmerkingen)
                    except AttributeError:
                        pass

                    db.add(new_init)
                    stats["created"] += 1
                    print(f"  + Created: [{excel_id}] {title[:50]}")

                db.commit()

            except Exception as e:
                db.rollback()
                stats["errors"] += 1
                print(f"  ✗ Error row {row_idx}: {e} — {title or '(geen titel)'}")

        print(f"\n{'='*50}")
        print(f"Import voltooid!")
        print(f"  Updated:  {stats['updated']}")
        print(f"  Created:  {stats['created']}")
        print(f"  Skipped:  {stats['skipped']}")
        print(f"  Errors:   {stats['errors']}")

    finally:
        db.close()


if __name__ == "__main__":
    excel_path = "20260806 Input/20260806 Inventarisatie_AI_initiatieven_v0.2.xlsx"
    if not Path(excel_path).exists():
        print(f"bestand niet gevonden: {excel_path}")
        sys.exit(1)
    import_initiatives(excel_path)
